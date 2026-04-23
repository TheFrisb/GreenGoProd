import mimetypes
import os
import re  # <--- NEW IMPORT
from collections import defaultdict
from datetime import datetime, timedelta

import openpyxl
from django.db.models import Case, CharField, Prefetch, Q, Value, When
from django.utils import timezone as django_timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Import your models
from .models import Order, OrderItem
from .utils import parse_offer_pieces

mimetypes.add_type("image/webp", ".webp")


class OrderExcelExporter:
    def __init__(self, date_from, date_to):
        self.date_from = date_from
        self.date_to = date_to
        self.timezone = django_timezone.get_current_timezone()

        self.wb = Workbook()
        self.ws = self.wb.active
        self.row_num = 1

        self.total_ordered_dict = defaultdict(int)
        self.total_ordered_stock_price = {}
        self.cart_offers_dict = defaultdict(int)
        self.upsell_offers_dict = defaultdict(int)
        self.thankyou_offers_dict = defaultdict(int)
        self.fee_offers_dict = defaultdict(int)
        self.duplicate_order_ids = set()
        self.same_products_duplicate_ids = set()

        self.thin_border = Side(border_style="thin", color="151515")
        self.border_all = Border(
            left=self.thin_border,
            right=self.thin_border,
            top=self.thin_border,
            bottom=self.thin_border,
        )
        self.header_font = Font(bold=True)

    def _get_stock_multiplier(self, attribute_name):
        """
        Returns N when attribute_name is an offer label like 'X2', 'X 3' or
        ' x4 ' (case-insensitive, whitespace-tolerant, full-string match).
        Returns 1 for sizes ('3 x 4m', '30cm x 60cm'), colors, or empty.
        """
        pieces = parse_offer_pieces(attribute_name)
        return pieces if pieces is not None else 1

    def _normalize_phone(self, phone):
        """
        Return the last 6 digits of `phone` (digits-only), or None if there
        is nothing to match on. Bridges Macedonian formats so that
        '+389 78 123 456', '078 123 456', '078-123-456', etc. all collapse
        to '123456'.
        """
        if not phone:
            return None
        digits = re.sub(r"\D", "", phone)
        if not digits:
            return None
        return digits[-6:]

    def _local_day_bounds(self, date_from, date_to):
        """Tz-aware half-open bounds covering local days [date_from, date_to]: use as gte/lt so the full local day of date_to is kept."""
        start_dt = django_timezone.make_aware(
            datetime.combine(date_from, datetime.min.time()),
            self.timezone,
        )
        end_dt = django_timezone.make_aware(
            datetime.combine(date_to + timedelta(days=1), datetime.min.time()),
            self.timezone,
        )
        return start_dt, end_dt

    def generate(self):
        self._setup_main_sheet()
        orders, duplicate_groups = self._get_orders_and_duplicates()

        for order in orders:
            count = order.id not in self.same_products_duplicate_ids
            self._write_order_row(order, count_in_summaries=count)

        self.row_num += 2
        self._write_duplicates_section(duplicate_groups)

        self.row_num += 4
        self._write_summaries()

        self._generate_procurement_sheet()

        return self.wb

    def _setup_main_sheet(self):
        widths = [20, 35, 35, 20, 20, 25, 10, 20, 65, 65, 10, 10, 100]
        columns = [
            "DATA NA PORACKA",
            "IME I PREZIME",
            "ADRESA",
            "GRAD",
            "TELEFON",
            "FEES",
            "VKUPNO",
            "DOSTAVA",
            "IME NA PRODUKT",
            "LABEL",
            "KOLICINA",
            "KOLICINA",
            "КОМЕНТАР",
        ]

        for i, width in enumerate(widths, 1):
            self.ws.column_dimensions[get_column_letter(i)].width = width
            self.ws.cell(row=self.row_num, column=i, value=columns[i - 1])

    def _get_orders_and_duplicates(self):
        base_qs = (
            Order.objects.filter(status__in=["Confirmed", "Pending"])
            .annotate(
                shippingann=Case(
                    When(shipping=True, then=Value("do vrata 190 den")),
                    When(shipping=False, then=Value("besplatna dostava")),
                    output_field=CharField(),
                )
            )
            .select_related()
            .prefetch_related(
                Prefetch(
                    "order",
                    queryset=OrderItem.objects.select_related(
                        "product", "product__supplier"
                    ),
                ),
                "orderfeesitem_set",
            )
        )

        # Walk a 10-day extended window in chronological order so the first
        # order from a given phone is always the "original" and any later
        # order from the same phone is flagged as a duplicate.
        lookback_start = self.date_from - timedelta(days=10)
        start_dt, end_dt = self._local_day_bounds(lookback_start, self.date_to)
        extended_orders = list(
            base_qs.filter(created_at__gte=start_dt, created_at__lt=end_dt)
            .order_by("created_at", "id")
        )

        def _skus_of(order):
            return {
                item.product.sku
                for item in order.order.all()
                if item.product is not None and item.product.sku
            }

        seen_originals = {}  # normalized phone -> first Order object
        duplicate_order_ids = set()
        dup_to_original = {}  # duplicate order id -> original Order object
        same_products_duplicate_ids = set()

        for order in extended_orders:
            norm = self._normalize_phone(order.number)
            if norm is None:
                # No usable phone -> never collides with anything.
                continue
            if norm not in seen_originals:
                seen_originals[norm] = order
            else:
                original = seen_originals[norm]
                duplicate_order_ids.add(order.id)
                dup_to_original[order.id] = original
                if _skus_of(order) == _skus_of(original):
                    same_products_duplicate_ids.add(order.id)

        self.duplicate_order_ids = duplicate_order_ids
        self.same_products_duplicate_ids = same_products_duplicate_ids

        # Main list: every in-range order (duplicates included). The
        # DUPLI NARACKI audit section below still lists the duplicates
        # grouped under their originals.
        current_orders = []
        in_range_duplicates = []
        for order in extended_orders:
            # Bucket by local-time date so orders created shortly after
            # local midnight aren't attributed to the previous UTC day.
            order_date = order.created_at.astimezone(self.timezone).date()
            if not (self.date_from <= order_date <= self.date_to):
                continue
            current_orders.append(order)
            if order.id in duplicate_order_ids:
                in_range_duplicates.append(order)

        current_orders.sort(key=lambda o: o.created_at, reverse=True)

        # Group each in-range duplicate under its first-encountered original
        # so the DUPLI NARACKI section can print the original as a header
        # before each cluster of duplicates (matches the pre-10-day-rewrite
        # display, where both sides of a duplicate pair were visible
        # together).
        groups_by_original_id = {}
        for dup in in_range_duplicates:
            original = dup_to_original[dup.id]
            if original.id not in groups_by_original_id:
                groups_by_original_id[original.id] = (original, [])
            groups_by_original_id[original.id][1].append(dup)

        duplicate_groups = list(groups_by_original_id.values())
        for _, dups in duplicate_groups:
            dups.sort(key=lambda o: o.created_at, reverse=True)
        duplicate_groups.sort(
            key=lambda g: max(
                [g[0].created_at] + [d.created_at for d in g[1]]
            ),
            reverse=True,
        )

        return current_orders, duplicate_groups

    def _write_order_row(self, order, count_in_summaries=True):
        self.row_num += 1
        order_items = order.order.all()
        order_fees = order.orderfeesitem_set.all()

        fees_text = ""
        items_name_text = ""
        items_label_text = ""
        total_quantity = 0
        is_priority = False

        for fee in order_fees:
            fees_text += f"{fee.title}\n"
            if count_in_summaries:
                self.fee_offers_dict[str(fee.title)] += 1
            if str(fee.title) in [
                "Приоритетна достава",
                "Приоритетна Достава + Осигурување на Пакет",
                "Бесплатна приоритетна достава",
            ]:
                is_priority = True

        processed_items_list = []
        for item in order_items:
            full_title = item.product.title
            if item.attribute_name:
                full_title += f" {item.attribute_name}"

            processed_items_list.append((full_title, item))

            if count_in_summaries:
                self.total_ordered_dict[full_title] += item.quantity

                # --- STOCK PRICE LOGIC (SHEET 1) ---
                multiplier = self._get_stock_multiplier(item.attribute_name)
                base_price = item.product.supplier_stock_price or 0
                self.total_ordered_stock_price[full_title] = base_price * multiplier
                # -----------------------------------

                if item.is_cart_offer:
                    if item.is_upsell_offer:
                        self.upsell_offers_dict[item.label] += item.quantity
                    else:
                        self.cart_offers_dict[item.label] += item.quantity
                if item.is_thankyou_offer:
                    self.thankyou_offers_dict[item.label] += item.quantity

            total_quantity += item.quantity

        written_titles = set()
        for full_title, item in processed_items_list:
            if full_title in written_titles:
                continue
            count = sum(1 for t, i in processed_items_list if t == full_title)
            qty_display = item.quantity + count - 1
            prefix = "PRIORITETNA " if is_priority else ""
            lbl = item.label if item.label else ""
            items_name_text += f"{prefix}{full_title} x {qty_display}"
            items_label_text += f"{prefix}{lbl} x {qty_display}"
            written_titles.add(full_title)

        height = 10 + (len(written_titles) * 15)
        self.ws.row_dimensions[self.row_num].height = height

        def w(col, val):
            c = self.ws.cell(row=self.row_num, column=col, value=val)
            c.alignment = Alignment(wrapText=True, vertical="top")

        date_str = order.created_at.astimezone(self.timezone).strftime(
            "%d.%m.%Y, %H:%M"
        )
        w(1, date_str)
        w(2, order.name)
        w(3, order.address)
        w(4, order.city)
        w(5, order.number)
        w(6, fees_text)
        w(7, order.total_price)
        w(8, order.shippingann)
        w(9, items_name_text)
        w(10, items_label_text)
        w(11, f"x{total_quantity}")
        w(12, total_quantity)
        w(13, order.message)

    def _write_duplicates_section(self, duplicate_groups):
        cell = self.ws.cell(
            row=self.row_num,
            column=2,
            value="DUPLI NARACKI" if duplicate_groups else "NEMA DUPLI NARACKI",
        )
        cell.font = self.header_font
        for original, dups in duplicate_groups:
            # Print the first-encountered order as the header row for this
            # duplicate cluster, then each duplicate beneath it. Neither
            # side mutates the Sheet 1 summary dicts: the original was
            # already counted in the main section above, and duplicates
            # must stay excluded from totals and Nabavki.
            self._write_order_row(original, count_in_summaries=False)
            for dup in dups:
                self._write_order_row(dup, count_in_summaries=False)

    def _write_summaries(self):
        def write_table(title, data_dict, include_price=False):
            self.ws.cell(row=self.row_num, column=9, value=title).font = (
                self.header_font
            )
            self.row_num += 1
            for key, value in data_dict.items():
                self.row_num += 1
                align = Alignment(wrapText=True, vertical="top", horizontal="left")
                c1 = self.ws.cell(row=self.row_num, column=9, value=str(key))
                c1.alignment = align
                c2 = self.ws.cell(row=self.row_num, column=10, value=value)
                c2.alignment = align
                if include_price:
                    c3 = self.ws.cell(
                        row=self.row_num,
                        column=11,
                        value=self.total_ordered_stock_price.get(key),
                    )
                    c3.alignment = align
            self.row_num += 2

        write_table("VKUPNA KOLICINA", self.total_ordered_dict, include_price=True)
        write_table("PONUDI VO KOSNICKA", self.cart_offers_dict)
        write_table("PONUDI NA PRODUCT PAGE", self.upsell_offers_dict)
        write_table("THANKYOU PONUDI", self.thankyou_offers_dict)
        write_table("CHECKOUT FEES", self.fee_offers_dict)

    def _generate_procurement_sheet(self):
        ws2 = self.wb.create_sheet("Nabavki")
        ws2.column_dimensions["A"].width = 20.3
        ws2.column_dimensions["B"].width = 30

        start_dt, end_dt = self._local_day_bounds(self.date_from, self.date_to)
        items = (
            OrderItem.objects.filter(
                Q(order__created_at__gte=start_dt),
                Q(order__created_at__lt=end_dt),
                Q(order__status__in=["Confirmed", "Pending"]),
            )
            .exclude(order_id__in=self.same_products_duplicate_ids)
            .select_related("product", "product__supplier")
        )

        grouped_data = defaultdict(lambda: defaultdict(lambda: {"qty": 0, "obj": None}))

        for item in items:
            supplier_name = item.product.supplier.name
            if item.attribute_name:
                label = f"{item.product.title} {item.attribute_name}"
            else:
                label = item.product.title

            entry = grouped_data[supplier_name][label]
            entry["qty"] += item.quantity
            if entry["obj"] is None:
                entry["obj"] = item

        sorted_suppliers = sorted(grouped_data.keys())
        row_num = 1

        for supplier_name in sorted_suppliers:
            ws2.row_dimensions[row_num].height = 30
            ws2.merge_cells(f"A{row_num}:E{row_num}")
            header = ws2.cell(row=row_num, column=1, value=supplier_name)
            header.font = Font(bold=True, size=24)
            header.alignment = Alignment(horizontal="center", vertical="center")
            for col in range(1, 6):
                ws2.cell(row=row_num, column=col).border = self.border_all

            row_num += 1
            start_row = row_num
            supplier_products = grouped_data[supplier_name]

            for label, data in supplier_products.items():
                qty = data["qty"]
                item_obj = data["obj"]
                product = item_obj.product

                ws2.row_dimensions[row_num].height = 113.5

                # Image (PNG)
                try:
                    if product.thumbnail:
                        image_path = product.export_image.path
                        if os.path.exists(image_path):
                            img = openpyxl.drawing.image.Image(image_path)
                            img.width = 150
                            img.height = 150
                            ws2.add_image(img, f"A{row_num}")
                except Exception as e:
                    print(f"Export Image Error for {label}: {e}")

                def cell(c, v, bold=False, fill=None):
                    cl = ws2.cell(row=row_num, column=c, value=v)
                    cl.alignment = Alignment(
                        wrapText=True, horizontal="center", vertical="center"
                    )
                    cl.border = self.border_all
                    if bold:
                        cl.font = Font(bold=True)
                    if fill:
                        cl.fill = fill
                    return cl

                # --- STOCK PRICE LOGIC (SHEET 2) ---
                multiplier = self._get_stock_multiplier(item_obj.attribute_name)
                base_price = product.supplier_stock_price or 0
                final_stock_price = base_price * multiplier
                # -----------------------------------

                cell(1, "")
                cell(2, label)
                cell(3, qty)
                cell(4, final_stock_price)  # Updated
                cell(5, f"=PRODUCT(C{row_num},D{row_num})")

                row_num += 1

            end_row = row_num
            yellow = PatternFill(
                start_color="FFFF00", end_color="FFFF00", patternType="solid"
            )

            sum_range_end = row_num - 1
            if sum_range_end >= start_row:
                total_cell = ws2.cell(
                    row=row_num, column=4, value=f"=SUM(E{start_row}:E{sum_range_end})"
                )
                total_cell.alignment = Alignment(horizontal="center", vertical="center")
                total_cell.border = self.border_all
                total_cell.font = Font(bold=True)
                total_cell.fill = yellow

            row_num += 3
