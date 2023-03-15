from django.shortcuts import render, redirect
from django.http import HttpRequest, HttpResponse, JsonResponse
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.views.generic import ListView, DetailView
from django.db.models import CharField, Value, Case, Value, When, Q, Func, F
from django.db.models import Q
from .models import *
from django.core.paginator import Paginator
from django.db.models.functions import Cast
import uuid
from uuid import  uuid4
import datetime
import xlwt
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.writer.excel import save_virtual_workbook
from django.db.models.functions import Concat
from .forms import ExportOrder
from django.utils.timezone import get_current_timezone, make_aware
import csv
from django.utils.html import strip_tags
from .controller import facebook_pixel
from openpyxl.drawing.image import Image
import openpyxl
from openpyxl.styles import Border, Side
from openpyxl.styles import PatternFill
# Create your views here.


def export_products_csv(request):
    products = Product.objects.all()

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="products.csv"'

    writer = csv.writer(response)
    writer.writerow(['ID', 'Title', 'Description', 'Link', 'Image Link', 'Availability', 'Price', 'Condition', 'Brand'])

    for product in products:
        content = strip_tags(product.content).replace('&nbsp;', '')
        content = os.linesep.join([s for s in content.splitlines() if s])
        if product.status == 'VARIABLE':
            attributes = ProductAttribute.objects.filter(product=product)
            for attribute in attributes:
                attribute_name = ''
                if(attribute.color is not None):
                    attribute_name = attribute.color.title

                if(attribute.size is not None):
                    attribute_name = attribute.size.title

                if(attribute.offer is not None):
                    attribute_name = attribute.offer.title
                
                writer.writerow([str(product.id) + '_' + attribute.label, product.title + ' - ' + attribute_name, content , 'https://greengoshop.mk' + product.get_absolute_url(), 'https://greengoshop.mk' + product.thumbnail.url, 'in stock', str(product.sale_price) + 'MKD', 'New', 'GreenGoShopMK'])
        
        writer.writerow([product.id, product.title, content , 'https://greengoshop.mk' + product.get_absolute_url(), 'https://greengoshop.mk' + product.thumbnail.url, 'in stock', str(product.sale_price) + 'MKD', 'New', 'GreenGoShopMK'])

    return response



def ProductListView(request):
    products = Product.objects.filter(status__in=['PUBLISHED','VARIABLE']).order_by('-date_posted')
    paginator = Paginator(products, 16)
    page = request.GET.get('page')
    products = paginator.get_page(page)
    total = 0
    context = {
        'products': products,
        'title': 'Почетна',
        }
    
    
    return render(request, 'shop/home.html', context)


def CategoryView(request, slug):
    if(Category.objects.filter(slug=slug)):
        products = Product.objects.filter(category__slug=slug).order_by('-date_posted')
        paginator = Paginator(products, 16)
        page = request.GET.get('page')
        products = paginator.get_page(page)
        title = Category.objects.get(slug = slug).name
        
        context = {
            'products': products, 
            'title': title,
            }
        
        return render(request, "shop/home.html", context)
    else:
        messages.warning(request, "Линкот што го следевте не постои")
        return redirect('shop-home')


def ProductView(request, slug):
    now = datetime.datetime.now()
    current_day = now.weekday()
    mapped_delivery_days = {
        0: ['Среда', 'Петок'],
        1: ['Четврток', 'Сабота'],
        2: ['Петок', 'Понеделник'],
        3: ['Сабота', 'Вторник'],
        4: ['Понеделник', 'Среда'],
        5: ['Вторник', 'Четврток'],
        6: ['Вторник', 'Четврток'],
    }
    delivery_days = mapped_delivery_days[current_day]
    try:
        attributes = ProductAttribute.objects.filter(product__slug=slug)
        gallery = ProductGallery.objects.filter(product__slug=slug)
        product = Product.objects.get(slug=slug)
        faq_toggle = ProductFAQ.objects.filter(product=product)
        if product.review_average != 0:
            reviews = Review.objects.filter(product__slug=slug)
        title = product.title
        percentage = 100 - int(product.sale_price / product.regular_price * 100)
        money_saved = product.regular_price - product.sale_price
        if(product.status != 'PRIVATE'):
            try:
                facebook_pixel.ViewContentEvent(request, product)
            except:
                pass
            
            if(product.review_average != 0):
                count = reviews.count()

                context = {
                    'product': product,
                    'reviews': reviews,
                    'reviewcount': count,
                    'slider1': Product.objects.filter(category__name='ЗАЛИХА')[:8],
                    'slider2': Product.objects.filter(status='PUBLISHED')[:8],
                    'gallery': gallery,
                    'attributes' : attributes,
                    'title': title,
                    'percentage': percentage,
                    'money_saved': money_saved,
                    'delivery_days': delivery_days,
                    'faq_toggle': faq_toggle,
                }
            else:
                context = {
                    'product': product,
                    'slider1': Product.objects.filter(category__name='ЗАЛИХА')[:8],
                    'slider2': Product.objects.filter(status='PUBLISHED')[:8],
                    'attributes' : attributes,
                    'gallery': gallery,
                    'title': title,
                    'percentage': percentage,
                    'money_saved': money_saved,
                    'delivery_days': delivery_days,
                    'faq_toggle': faq_toggle,
                }


            return render(request, "shop/product-page.html", context)
        else:
            return redirect('shop-home')
    except:
        messages.warning(request, "Линкот што го следевте не постои")
        return redirect('shop-home')


def CheckoutView(request):
    orderFees = CheckoutFees.objects.all()
    try:
        facebook_pixel.InitiateCheckoutEvent(request)
    except:
        pass
    try:
        cartFees = CartFees.objects.filter(cart__session=request.session['nonuser'])
    except:
        cartFees = None
    
    feetotal = 0
    for orderfee in orderFees:
        for cartfee in cartFees or []:
            if(orderfee == cartfee.fee):
                orderfee.is_added = True
                feetotal += cartfee.price
    
    context = {
        'title': 'Кон Нарачка',
        'orderFees': orderFees,
        'cartFees': cartFees,
        'feetotal': feetotal,
        }

    return render(request, 'shop/checkout.html', context)


def ThankYouView(request, slug):
    order = Order.objects.filter(tracking_no=slug).first
    if(order):   
        orderItems = OrderItem.objects.filter(order__tracking_no=slug) # Sql join ?
        offerproduct = orderItems.reverse()[0]
        if offerproduct.attribute_price is not None:
            offerproduct.attribute_price = offerproduct.attribute_price - offerproduct.attribute_price * 20 // 100 
        else:
            offerproduct.price = offerproduct.price - offerproduct.price * 20 // 100 
        orderFees = OrderFeesItem.objects.filter(order__tracking_no=slug)
        feetotal = 0
        for fee in orderFees:
            feetotal += fee.price
    title = 'Ви благодариме!'
    context = {
        'order': order,
        'orderItems': orderItems,
        'offerproduct': offerproduct,
        'orderFees': orderFees,
        'feetotal': feetotal,
        'title': title,
    }

    return render(request, 'shop/thank-you.html', context)




class SearchResultsView(ListView):
    model = Product
    template_name = 'shop/home.html'
    extra_context = {
        'title' : "Барање"
    }
    context_object_name = 'products'
    def get_queryset(self):
        query = self.request.GET.get("q")
        object_list = Product.objects.filter(Q(title__icontains=query, status__in=['PUBLISHED','VARIABLE']) | Q(sku__icontains=query, status__in=['PUBLISHED','VARIABLE']))
        
        return object_list
    



def login_shopmanager(request):
    if request.user.is_authenticated:
        messages.warning(request, "You are already logged in")
        return redirect("shopmanagerhome")
    else:
        context = {
                'title': 'Login',
            }
        if request.method =='POST':
            name = request.POST.get('username')
            password = request.POST.get('password')
            
            user = authenticate(request, username=name, password=password)

            if user is not None:
                login(request, user)
                messages.success(request, "Logged in Successfully")
                return redirect("shopmanagerhome")
            else:
                messages.error(request, "Invalid username or Password")
                return redirect('/')
        return render(request, "shop/shopmanager/login.html", context)


def logout_shopmanager(request):
    if request.user.is_authenticated:
        logout(request)
        messages.success(request, "Logged out successfully")
    
    return redirect("/")
    


def shopmanager_dashboard(request):
    orders = Order.objects.filter(status='Pending').order_by('-id')[:50]
    orderItems = OrderItem.objects.filter(order__status = 'Pending').order_by('-id')
    orderfees = OrderFeesItem.objects.filter(order__status = 'Pending').order_by('-id')
    title = 'НЕПОТВРДЕНИ НАРАЧКИ'
    form = ExportOrder()
    context = {
        'orders' : orders,
        'orderItems': orderItems,
        'orderFees': orderfees,
        'heading' : title,
        'order_status': 'Непотврдена',
        'form': form,
    }
    return render(request, "shop/shopmanager/dashboard.html", context)


def shopmanager_confirmed(request):
    orders = Order.objects.filter(status='Confirmed').order_by('-updated_at')[:50]
    orderItems = OrderItem.objects.filter(order__status = 'Confirmed').order_by('-id')
    orderfees = OrderFeesItem.objects.filter(order__status = 'Confirmed').order_by('-id')
    title = 'ПОТВРДЕНИ НАРАЧКИ'
    context = {
        'orders' : orders,
        'orderItems': orderItems,
        'orderFees': orderfees,
        'heading' : title,
        'order_status': 'Потврдена',
    }
    return render(request, "shop/shopmanager/dashboard.html", context)


def shopmanager_abandoned_carts(request):
    abandoned_carts = Abandoned_Carts.objects.all().order_by('-id')
    abandoned_cartItems = Abandoned_Carts.objects.all().order_by('-id')
    paginator = Paginator(abandoned_carts, 50)
    page = request.GET.get('page')
    carts = paginator.get_page(page)
    #print(abandoned_carts.first().get_items)
    context = {
        'carts': carts,
        'cartItems': abandoned_cartItems,
        'title': 'Abandoned Carts',
    }
    return render(request, "shop/shopmanager/abandoned_carts.html", context)



def shopmanager_deleted(request):
    orders = Order.objects.filter(status='Deleted').order_by('-updated_at')[:50]
    orderItems = OrderItem.objects.filter(order__status = 'Deleted').order_by('-id')
    orderfees = OrderFeesItem.objects.filter(order__status = 'Deleted').order_by('-id')
    title = 'ИЗБРИШЕНИ НАРАЧКИ'
    context = {
        'orders' : orders,
        'orderItems': orderItems,
        'orderFees': orderfees,
        'heading' : title,
        'order_status': 'Избришена',
    }
    return render(request, "shop/shopmanager/dashboard.html", context)


def shopmanager_create_order(request):
    context = {
        'heading': 'Креири нарачка'
    }

    return render(request, 'shop/shopmanager/create_order.html', context)

def Dostava(request):
    context = {
        'title': 'Политика за достава'
    }
    return render(request, 'shop/policies/dostava.html', context)


def Reklamacija(request):
    context = {
        'title': 'Политика за рекламација'
    }
    return render(request, 'shop/policies/reklamacija.html', context)


def Pravila_Na_Koristenje(request):
    context = {
        'title': 'Правила на користење'
    }
    return render(request, 'shop/policies/pravila_na_koristenje.html', context)

def Cookies_Page(request):
    context = {
        'title': 'Политика за приватност и колачиња'
    }
    return render(request, 'shop/policies/politika_na_privatnost_i_kolacinja.html', context)


def export_excel(request):
    if request.method == 'POST':
        form = ExportOrder(request.POST)
        if form.is_valid():
            timezone = get_current_timezone()
            date_from = form.cleaned_data["date_from"]
            date_to = form.cleaned_data["date_to"]
            total_ordered_dict = {}
            cart_offers_dict = {}
            thankyou_offers_dict = {}
            workbook = Workbook()
            worksheet = workbook.active
            row_num = 1
            
            worksheet.column_dimensions['A'].width = 20
            worksheet.column_dimensions['B'].width = 35
            worksheet.column_dimensions['C'].width = 35
            worksheet.column_dimensions['D'].width = 20
            worksheet.column_dimensions['E'].width = 20
            worksheet.column_dimensions['F'].width = 25
            worksheet.column_dimensions['G'].width = 10
            worksheet.column_dimensions['H'].width = 20
            worksheet.column_dimensions['I'].width = 65
            worksheet.column_dimensions['J'].width = 65
            worksheet.column_dimensions['K'].width = 10
            worksheet.column_dimensions['L'].width = 10
            worksheet.column_dimensions['M'].width = 100
            columns = ['DATA NA PORACKA', 'IME I PREZIME', 'ADRESA', 'GRAD', 'TELEFON', 'FEES', 'VKUPNO', 'DOSTAVA', 'IME NA PRODUKT', 'LABEL','KOLICINA', 'KOLICINA', 'КОМЕНТАР']
            for col_num in range(1, len(columns)+1):             
                cell =  worksheet.cell(row=row_num, column=col_num, value=columns[col_num - 1])

            rows = Order.objects.filter(Q(created_at__range = [date_from, date_to], status = 'Confirmed',) | Q(created_at__range = [date_from, date_to], status = 'Pending')).annotate(
                shippingann = Case(
                    When(shipping = True, then=Value('do vrata 130 den')),
                    When(shipping = False, then=Value('besplatna dostava'))
                ),
            ).order_by('-created_at').values_list('created_at', 'name', 'address', 'city', 'number', 'tracking_no', 'total_price', 'shippingann', 'number', 'number', 'number', 'number', 'message')
            print(rows)
            total_ordered_stock_price = {}
            for row in rows:
                row_num += 1
                height = 10
                height2 = 10
                order_items = OrderItem.objects.filter(order__tracking_no = row[5]).annotate(full_product_title = Concat('product__title', Value(' '), 'attribute_name' ))
                order_fees = OrderFeesItem.objects.filter(order__tracking_no = row[5])
                order_items_total_name = ''
                order_items_total_label = ''
                order_fees_total = ''
                quantity = 0
                priority = False

                for fee in order_fees:
                    order_fees_total += str(fee.title) + '\n'
                    if(str(fee.title) == 'Приоритетна достава'):
                        priority = True
                    height2 +=15
                    
                for item in order_items:
                    quantity += item.quantity
                    if item.label in total_ordered_dict:
                        total_ordered_dict[item.label] += item.quantity
                    else:
                        total_ordered_dict[item.label] = item.quantity
                        total_ordered_stock_price[item.label] = item.product.supplier_stock_price
                        
                for item in order_items:
                    if item.is_cart_offer is True:
                        if item.label in cart_offers_dict:
                            cart_offers_dict[item.label] += item.quantity
                        else:
                            cart_offers_dict[item.label] = item.quantity

                    if item.is_thankyou_offer is True:
                        if item.label in thankyou_offers_dict:
                            thankyou_offers_dict[item.label] += item.quantity
                        else:
                            thankyou_offers_dict[item.label] = item.quantity
                            

                for item in order_items:
                    occurence = 0
                    
                    for item_check in order_items:
                        if item_check.full_product_title == item.full_product_title:
                            occurence += 1
                            if occurence > 1:
                                item_check.full_product_title = ''
                                item_check.label = ''

                    if item.full_product_title!='':
                        if priority is True:
                            order_items_total_name += 'PRIORITETNA ' + str(item.full_product_title) + ' x ' + str(item.quantity + occurence - 1)
                            order_items_total_label += 'PRIORITETNA ' + str(item.label) + ' x ' + str(item.quantity + occurence - 1)
                        else:
                            order_items_total_name += str(item.full_product_title) + ' x ' + str(item.quantity + occurence - 1)
                            order_items_total_label += str(item.label) + ' x ' + str(item.quantity + occurence - 1)
                        height += 15


               # if(height2 > height):
                  #  height = height2

                worksheet.row_dimensions[row_num].height = height
                for col_num in range(1, len(row)+1):
                    worksheet.cell(row=row_num, column=col_num).alignment = Alignment(wrapText=True,  vertical='top')

                    if(col_num == 1):
                        date = row[col_num-1].astimezone(timezone)      
                        cell =  worksheet.cell(row=row_num, column=col_num).value = (date.strftime("%d.%m.%Y, %H:%M"))
                    elif(col_num == 6):
                        cell =  worksheet.cell(row=row_num, column=col_num).value = order_fees_total
                    elif(col_num == 9):
                        cell =  worksheet.cell(row=row_num, column=col_num).value = order_items_total_name
                    elif(col_num == 10):
                        cell =  worksheet.cell(row=row_num, column=col_num).value = order_items_total_label
                    elif(col_num == 11):
                        cell =  worksheet.cell(row=row_num, column=col_num).value = 'x' + str(quantity)
                    elif(col_num == 12):
                        cell =  worksheet.cell(row=row_num, column=col_num).value = quantity
                            
                    else:
                        cell =  worksheet.cell(row=row_num, column=col_num).value = str(row[col_num-1])

            print(total_ordered_dict)
            row_num += 4
            worksheet.cell(row=row_num, column=9).font = Font(bold=True)
            cell = worksheet.cell(row=row_num, column=9).value = 'VKUPNA KOLICINA'
            row_num += 1
            for key, value in total_ordered_dict.items():
                row_num += 1
                i = 0
                
                worksheet.cell(row=row_num, column=9).alignment = Alignment(wrapText=True,  vertical='top', horizontal='left')
                worksheet.cell(row=row_num, column=10).alignment = Alignment(wrapText=True,  vertical='top',horizontal='left')
                worksheet.cell(row=row_num, column=11).alignment = Alignment(wrapText=True,  vertical='top',horizontal='left')
                cell = worksheet.cell(row=row_num, column=9).value = str(key)
                cell = worksheet.cell(row=row_num, column = 10).value = value
                cell = worksheet.cell(row=row_num, column = 11). value = total_ordered_stock_price.get(key)
                i += 1
            
            row_num += 2
            worksheet.cell(row=row_num, column=9).font = Font(bold=True)
            cell = worksheet.cell(row=row_num, column=9).value = 'PONUDI VO KOSNICKA'
            for key, value in cart_offers_dict.items():
                row_num += 1
                i = 0
                
                worksheet.cell(row=row_num, column=9).alignment = Alignment(wrapText=True,  vertical='top', horizontal='left')
                worksheet.cell(row=row_num, column=10).alignment = Alignment(wrapText=True,  vertical='top',horizontal='left')
                cell = worksheet.cell(row=row_num, column=9).value = str(key)
                cell = worksheet.cell(row=row_num, column = 10).value = value
                i += 1

            row_num += 2
            worksheet.cell(row=row_num, column=9).font = Font(bold=True)
            cell = worksheet.cell(row=row_num, column=9).value = 'THANKYOU PONUDI'
            for key, value in thankyou_offers_dict.items():
                row_num += 1
                i = 0
                
                worksheet.cell(row=row_num, column=9).alignment = Alignment(wrapText=True,  vertical='top', horizontal='left')
                worksheet.cell(row=row_num, column=10).alignment = Alignment(wrapText=True,  vertical='top',horizontal='left')
                cell = worksheet.cell(row=row_num, column=9).value = str(key)
                cell = worksheet.cell(row=row_num, column = 10).value = value
                i += 1
            nabavki_dict = {}
            nabavki_list = []
            rows2 = OrderItem.objects.filter(Q(order__created_at__range = [date_from, date_to], order__status = 'Confirmed',) | Q(order__created_at__range = [date_from, date_to], order__status = 'Pending')).all()
            
            for row in rows2:
                
                label = row.label
                quantity = row.quantity
                stock_price = row.product.supplier_stock_price
                thumbnail_url = row.product.export_image.path
                supplier = row.product.supplier.name

                if label in nabavki_dict:
                    nabavki_dict[label]["quantity"] += quantity
                else:
                    nabavki_dict[label] = {"supplier": supplier, "stock_price": stock_price, "thumbnail_url": thumbnail_url, "quantity": quantity}

            nabavki_list = [{"label": key, "supplier": value["supplier"], "stock_price": value["stock_price"], "thumbnail_url": value["thumbnail_url"], "quantity": value["quantity"]} for key, value in nabavki_dict.items()]
            row_num2 = 1
            worksheet2 = workbook.create_sheet("Nabavki")
            worksheet2.column_dimensions['A'].width = 20.3
            worksheet2.column_dimensions['B'].width = 30

            suppliers = Dobavuvac.objects.all()
            for supplier in suppliers:
                name = supplier.name
                worksheet2.row_dimensions[row_num2].height = 30
                worksheet2.cell(row=row_num2, column=1).alignment = Alignment(horizontal='center', vertical='center')
                worksheet2.cell(row=row_num2, column=1).font = Font(bold=True, size=24)
                thin = Side(border_style='thin', color='151515')
                worksheet2.cell(row=row_num2, column=1).border = Border(left=thin, right=thin, top=thin, bottom=thin)
                worksheet2.cell(row=row_num2, column=5).border = Border(right=thin)
                worksheet2.merge_cells("A" + str(row_num2) + ":E" + str(row_num2))
                worksheet2.cell(row=row_num2, column=1).value = name
                row_num2 += 1
                start_row = row_num2
                total = 0
                for item in nabavki_list:
                    if item["supplier"] == name:
                        worksheet2.row_dimensions[row_num2].height = 113.5
                        if "thumbnail_url" in item:
                            img = openpyxl.drawing.image.Image(item["thumbnail_url"])
                            worksheet2.add_image(img, 'A' + str(row_num2))

                        worksheet2.cell(row=row_num2, column=2).alignment = Alignment(wrapText=True,  horizontal='center', vertical='center')
                        worksheet2.cell(row=row_num2, column=3).alignment = Alignment(wrapText=True,  horizontal='center', vertical='center')
                        worksheet2.cell(row=row_num2, column=4).alignment = Alignment(wrapText=True,  horizontal='center', vertical='center')
                        worksheet2.cell(row=row_num2, column=5).alignment = Alignment(horizontal='center', vertical='center')
                        worksheet2.cell(row=row_num2, column=1).alignment = Alignment(horizontal='center', vertical='center')
                        worksheet2.cell(row=row_num2, column=2).border = Border(left=thin, right=thin, top=thin, bottom=thin)
                        worksheet2.cell(row=row_num2, column=3).border = Border(left=thin, right=thin, top=thin, bottom=thin)
                        worksheet2.cell(row=row_num2, column=4).border = Border(left=thin, right=thin, top=thin, bottom=thin)
                        worksheet2.cell(row=row_num2, column=1).border = Border(left=thin, right=thin, top=thin, bottom=thin)
                        worksheet2.cell(row=row_num2, column=5).border = Border(left=thin, right=thin, top=thin, bottom=thin)
                        worksheet2.cell(row=row_num2, column=2).value = item["label"]
                        worksheet2.cell(row=row_num2, column=3).value = item["quantity"]
                        worksheet2.cell(row=row_num2, column=4).value = item["stock_price"]
                        worksheet2.cell(row=row_num2, column=5).value = f'=PRODUCT(C{row_num2},D{row_num2})'
                        row_num2 += 1
                end_row = row_num2
                fill = PatternFill(start_color='FFFF00', end_color='FFFF00', patternType='solid')
                worksheet2.cell(row=row_num2, column=4).alignment = Alignment(horizontal='center', vertical='center')
                worksheet2.cell(row=row_num2, column=4).border = Border(left=thin, right=thin, top=thin, bottom=thin)
                worksheet2.cell(row=row_num2, column=4).font = Font(bold=True)
                worksheet2.cell(row=row_num2, column=4).fill = fill
                worksheet2.cell(row=row_num2, column=4).value = f'=SUM(E{start_row}:E{end_row})'


                row_num2 += 3

            response = HttpResponse(content=save_virtual_workbook(workbook))
            
            response['Content-Disposition'] = 'attachment; filename=eksport_' + str(date_from) + ' - ' + str(date_to) + '.xlsx'
            return response
    
    return redirect('/')
  
    
def get_recent_ordered(request):
    if request.method == 'GET':
        random_int = randint(1, 20)
        product = Product.objects.order_by('-pk').filter(status='PUBLISHED')[random_int]
        return JsonResponse({
            'url': product.get_absolute_url(),
            'thumbnail': product.thumbnail_loop.url,
            'title': product.title,
            'regular_price': product.regular_price,
            'sale_price': product.sale_price,
            })
    else:
        return redirect('/')
  
def create_or_check_abandoned_cart(request):
    if request.method == 'POST':
        current_cart = Cart.objects.get(session = request.session['nonuser'])
        current_cartitems = CartItems.objects.filter(cart=current_cart)

        if current_cartitems.count() == 0:
            return JsonResponse({'status': "No cartItems"})
        
        else:     
            abandoned_name = str(request.POST.get('name'))
            abandoned_phone_number = str(request.POST.get('phone'))
            abandoned_address = str(request.POST.get('address'))
            # abandoned_cart = Abandoned_Carts.objects.update_or_create(session = request.session['nonuser'], defaults={
            #     'name': abandoned_name,
            #     'phone': abandoned_phone_number
            # })
            
            try:
                abandoned_cart = Abandoned_Carts.objects.get(session = request.session['nonuser'])
                abandoned_cart.name = abandoned_name
                abandoned_cart.phone = abandoned_phone_number
                abandoned_cart.address = abandoned_address
                abandoned_cart.save()
                print('Found and updated')
            except:
                abandoned_cart = Abandoned_Carts.objects.create(session = request.session['nonuser'])
                abandoned_cart.name = abandoned_name
                abandoned_cart.phone = abandoned_phone_number
                abandoned_cart.address = abandoned_address
                abandoned_cart.save()
                print('Not Found and created')
            print(abandoned_cart)

            for current_item in current_cartitems:
                abandoned_item = Abandoned_CartItems.objects.filter(cart=abandoned_cart, product = current_item.product,
                                                                    attributename=current_item.attributename, product_qty = current_item.product_qty,
                                                                    attribute=current_item.attribute, attributeprice=current_item.attributeprice,
                                                                    offer_price=current_item.offer_price).first()
                print('Abandoned item: ', abandoned_item)
                if not abandoned_item:
                    Abandoned_CartItems.objects.create(cart=abandoned_cart, product = current_item.product,
                                                                    attributename=current_item.attributename, product_qty = current_item.product_qty,
                                                                    attribute=current_item.attribute, attributeprice=current_item.attributeprice,
                                                                    offer_price=current_item.offer_price)

        return JsonResponse({'status': "Success"})
    else:
        return redirect('/')
    


def remove_abandoned_cart(request):
    if request.method == 'POST':
        id = str(request.POST.get('cartId'))
        Abandoned_Carts.objects.get(pk=id).delete()
        return JsonResponse({'status': "Success"})
    else:
        return redirect('/')
    
